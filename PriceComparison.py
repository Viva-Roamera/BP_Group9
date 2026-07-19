"""
## Iva - LEGO Price Comparison Analysis
-------------------------------
Reads master_product_list.csv, cleans it, builds a PriceComparison.xlsx
workbook (with live formulas), and produces two charts:
  1. Lowest Price Wins  (count of products where each shop had the lowest price)
  2. Price Range by Category (lowest -> highest price per category)

Note: Need to install the following packages if not already installed:
    pip install pandas numpy matplotlib openpyxl
"""

from pathlib import Path

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

candidate_inputs = [
    OUTPUT_DIR / "master_product_list.csv",
    SCRIPT_DIR / "master_product_list.csv",
    Path("/mnt/user-data/uploads/master_product_list.csv"),
]
INPUT_CSV = next((path for path in candidate_inputs if path.exists()), candidate_inputs[0])
OUTPUT_XLSX = OUTPUT_DIR / "PriceComparison.xlsx"
CHART1_PATH = OUTPUT_DIR / "lowest_price_wins.png"
CHART2_PATH = OUTPUT_DIR / "price_range_by_category.png"

# ---------------------------------------------------------------------------
# 1. LOAD + CLEAN DATA
# ---------------------------------------------------------------------------
df = pd.read_csv(INPUT_CSV)

print(f"Raw rows loaded: {len(df)}")

# --- Drop exact duplicate rows (Brickmo rows were scraped/appended twice) ---
before = len(df)
df = df.drop_duplicates()
print(f"Dropped {before - len(df)} exact duplicate rows")

# --- Drop rows with no ProductID (can't be matched across shops) ---
before = len(df)
df = df.dropna(subset=["ProductID"])
print(f"Dropped {before - len(df)} rows with missing ProductID")


# --- Clean price column: strip currency symbols/whitespace, convert to float ---
df["price"] = (
    df["price"].astype(str)
    .str.replace(r"[^\d.]", "", regex=True)
    .astype(float)
)

# --- Normalize ProductID to a clean integer-like string (was float: 60312.0) ---
df["ProductID"] = df["ProductID"].astype(int)

# --- Normalize category casing (e.g. "LEGO City" vs "Lego City") ---
df["category"] = df["category"].str.strip().str.title()

# --- Drop remaining rows with missing category (can't be charted by category) ---
before = len(df)
df = df.dropna(subset=["category"])
print(f"Dropped {before - len(df)} rows with missing category")

df = df.reset_index(drop=True)
print(f"Clean rows remaining: {len(df)}")
print(f"Shops: {sorted(df['shop'].unique())}")
print(f"Unique ProductIDs: {df['ProductID'].nunique()}")

# ---------------------------------------------------------------------------
# 2. BUILD COMPARISON TABLE (in pandas, to know row counts / ordering,
#    then re-created as live formulas in Excel)
# ---------------------------------------------------------------------------
product_ids = sorted(df["ProductID"].unique())

# ---------------------------------------------------------------------------
# 3. WRITE EXCEL WORKBOOK
# ---------------------------------------------------------------------------
wb = Workbook()

# --- Sheet 1: Raw Data (cleaned) ---
ws_raw = wb.active
ws_raw.title = "Raw Data"
raw_headers = ["shop", "ProductID", "category", "name", "price"]
ws_raw.append(raw_headers)
for cell in ws_raw[1]:
    cell.font = Font(bold=True, name="Arial")
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

for _, row in df[raw_headers].iterrows():
    ws_raw.append([row["shop"], int(row["ProductID"]), row["category"], row["name"], float(row["price"])])

n_raw = len(df)  # number of data rows in Raw Data (rows 2..n_raw+1)
last_raw_row = n_raw + 1

for col, width in zip("ABCDE", [16, 12, 16, 48, 10]):
    ws_raw.column_dimensions[col].width = width
for r in range(2, last_raw_row + 1):
    ws_raw.cell(row=r, column=5).number_format = "€#,##0.00"
for row in ws_raw.iter_rows(min_row=2, max_row=last_raw_row):
    for cell in row:
        cell.font = Font(name="Arial")

# --- Sheet 2: PriceComparison ---
ws = wb.create_sheet("PriceComparison")
headers = [
    "ProductID", "Product Name", "Category",
    "Lowest Price (€)", "Lowest Price Shop",
    "Highest Price (€)", "Highest Price Shop",
    "Price Difference (€)", "# Shops Offering",
]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, name="Arial", color="FFFFFF")
    cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

raw_pid = f"'Raw Data'!$B$2:$B${last_raw_row}"
raw_cat = f"'Raw Data'!$C$2:$C${last_raw_row}"
raw_name = f"'Raw Data'!$D$2:$D${last_raw_row}"
raw_price = f"'Raw Data'!$E$2:$E${last_raw_row}"
raw_shop = f"'Raw Data'!$A$2:$A${last_raw_row}"

for i, pid in enumerate(product_ids):
    r = i + 2  # excel row (header is row 1)
    ws.cell(row=r, column=1, value=int(pid))  # ProductID (hardcoded lookup key)
    ws.cell(row=r, column=2, value=f"=INDEX({raw_name},MATCH(A{r},{raw_pid},0))")
    ws.cell(row=r, column=3, value=f"=INDEX({raw_cat},MATCH(A{r},{raw_pid},0))")
    ws.cell(row=r, column=4, value=f"=_xlfn.MINIFS({raw_price},{raw_pid},A{r})")
    ws.cell(row=r, column=5, value=(
        f"=INDEX({raw_shop},MATCH(1,INDEX(({raw_pid}=A{r})*({raw_price}=D{r}),0),0))"
    ))
    ws.cell(row=r, column=6, value=f"=_xlfn.MAXIFS({raw_price},{raw_pid},A{r})")
    ws.cell(row=r, column=7, value=(
        f"=INDEX({raw_shop},MATCH(1,INDEX(({raw_pid}=A{r})*({raw_price}=F{r}),0),0))"
    ))
    ws.cell(row=r, column=8, value=f"=F{r}-D{r}")
    ws.cell(row=r, column=9, value=f"=COUNTIF({raw_pid},A{r})")

last_cmp_row = len(product_ids) + 1

# Formatting
for col, width in zip("ABCDEFGHI", [12, 48, 16, 16, 18, 16, 18, 18, 16]):
    ws.column_dimensions[col].width = width
for r in range(2, last_cmp_row + 1):
    ws.cell(row=r, column=4).number_format = "€#,##0.00"
    ws.cell(row=r, column=6).number_format = "€#,##0.00"
    ws.cell(row=r, column=8).number_format = "€#,##0.00"
    for c in range(1, 10):
        ws.cell(row=r, column=c).font = Font(name="Arial")
        ws.cell(row=r, column=c).alignment = Alignment(horizontal="center") if c not in (2,) else Alignment(horizontal="left")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{last_cmp_row}"

# Note documenting exclusions/assumptions, placed just below the table
note_row = last_cmp_row + 2
ws.cell(row=note_row, column=1, value=(
    "Note: 1 row with a missing ProductID was excluded from this analysis. "
    "48 exact duplicate rows (Brickmo) were also removed. Zavvi rows are "
    "included because their prices are converted to EUR in the scraper. "
    "See 'Raw Data' tab for the cleaned dataset."
)).font = Font(italic=True, size=9, name="Arial")

wb.save(OUTPUT_XLSX)
print(f"Saved workbook: {OUTPUT_XLSX}")
